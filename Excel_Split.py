import os
import pandas as pd
from openpyxl import Workbook
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Load the main Excel file
current_directory = os.path.dirname(os.path.abspath(__file__))
input_file = os.path.join(current_directory, "Excel_input/Master.xlsx")


def opening():
    global input_file, current_directory
    folder_path = os.path.join(current_directory, "Excel_output")
    if os.listdir(folder_path):
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)
            except Exception as e:
                print(f'Failed to delete {file_path}. Reason: {e}')

    df = pd.read_excel(input_file)
    grouped = df.groupby("Employee ID")
    count = len(grouped)
    return count


def split():
    global input_file
    # input_file = "Excel_input/Master.xlsx"
    df = pd.read_excel(input_file)
    output = ''
    # Group the data by Employee ID
    grouped = df.groupby("Employee ID")

    # Iterate through each group
    for employee_id, group in grouped:
        # Create a new workbook for each employee
        wb = Workbook()
        ws = wb.active

        # Add the headers in the first row
        headers = df.columns.tolist()
        ws.append(headers)

        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

        header_font = Font(bold=True, color="FFFFFF")  # White text, Bold
        total_font = Font(bold=True, color="000000")  # White text, Bold
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue fill
        total_fill = PatternFill(start_color="fff2cc", end_color="fff2cc", fill_type="solid")  # Blue fill
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        alignment_center = Alignment(horizontal="center", vertical="center")

        # Skip the second row (leaving it blank)
        # ws.append([])

        # Add the Employee ID and Name in the third row
        employee_row = group.iloc[0][["Employee ID", "Name"]].values.tolist()
        # ws.append(employee_row)

        # Prepare date-related data
        date_columns = [col for col in group.columns if col not in ["Employee ID", "Name"]]
        rows_to_add = []
        sum_row = {}
        print(date_columns)

        for _, row in group.iterrows():
            # Split date-related columns and handle multiple rows
            max_splits = 1
            split_data = {col: [] for col in date_columns}
            print(split_data)

            for col in date_columns:
                print("Col ", col)
                if pd.isna(row[col]):
                    split_data[col] = [""]
                else:
                    split_values = [x.strip() for x in str(row[col]).split(",")]
                    print("spl ", len(split_values))
                    if len(split_values) < 2:
                        split_values = [datetime.strptime(split_values[0], "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y")]

                    split_data[col] = split_values
                    max_splits = max(max_splits, len(split_values))
            print(max_splits)
            print(split_data)

            # for col in date_columns:
            #     sums = len(split_data[col])

            sum_row = {col: (len(split_data[col]) if split_data[col][0] != "" else "") for col in date_columns}
            # print(sums)
            print(sum_row)
            rows_to_add.append(sum_row)
            # Construct rows based on splits
            for i in range(max_splits):
                new_row = {col: (split_data[col][i] if i < len(split_data[col]) else "") for col in date_columns}
                print("new r ", new_row)
                rows_to_add.append(new_row)
            print(rows_to_add)
            print(len(rows_to_add))
        # Add the split rows starting from the third row (same row as Employee ID and Name)
        for idx, row in enumerate(rows_to_add):
            print("ind", idx)
            print(row)
            if idx == 0:
                # ws.cell(row=1, column=1, value="Total")
                # ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=1)
                for col_idx, col in enumerate(date_columns, start=3):  # Start from Column 3
                    ws.cell(row=2, column=col_idx, value=row.get(col, ""))
            elif idx == 1:
                # Merge date data with the Employee ID and Name for the first row
                ws.append([employee_row[0], employee_row[1]] + [row.get(col, "") for col in date_columns])
            else:
                # Add only date data for subsequent rows
                ws.append(["", ""] + [row.get(col, "") for col in date_columns])

        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)  # Get column letter (e.g., A, B, C)

            for cell in col:
                try:
                    # Measure cell content length
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = max_length + 2  # Add some extra padding
            ws.column_dimensions[column_letter].width = adjusted_width

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = alignment_center

        for col_idx, cell in enumerate(ws[1], start=1):  # ws[1] represents the first row
            cell.font = header_font
            cell.fill = header_fill

        for col_idx, cell in enumerate(ws[2], start=3):
            cell.font = total_font
            cell.fill = total_fill

        # Save the individual file

        output_file = os.path.join(current_directory, f"Excel_output/Employee_{employee_id}.xlsx")
        wb.save(output_file)
        print(f"The Created file: {output_file}")

        output += f"The Created file: Employee_{employee_id}.xlsx"
    return output


# opening()
